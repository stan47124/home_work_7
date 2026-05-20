import io
import os
import zipfile
import csv
from openpyxl import load_workbook
from pypdf import PdfReader

# Прописываем путь для архива

current_dir = os.path.dirname(os.path.abspath(__file__))
results_dir = os.path.join(current_dir, "results")

archive_path = os.path.join(results_dir, "archive.zip")


# Создание архива

with zipfile.ZipFile(archive_path, 'w') as zipf:
    zipf.write('test/resources/test_file_1.csv')
    zipf.write('test/resources/test_file_2.pdf')
    zipf.write('test/resources/test_file_3.xlsx')

# Проверка, что архив создан

def archive_exists(archive_path):
    assert os.path.isfile(archive_path), f"Архив не найден"

# Проверка количества файлов

def amount_of_files(archive_path):
    with zipfile.ZipFile(archive_path, 'r') as zipf:
        assert len(zipf.namelist()) == 3, f"Количество файлов не равно 3"

# Проверка на уникальность

def unic_files(archive_path):
    with zipfile.ZipFile(archive_path, 'r') as zipf:
        seen = set()
        for item in zipf.namelist():
            if item in seen:
                assert False, f"Найден дубликат: {item}"
            else:
                seen.add(item)

# Проверка на пустоту и битые файлы

def files_are_not_empty(archive_path):
    with zipfile.ZipFile(archive_path, 'r') as zipf:
        for filename in zipf.namelist():
            assert len(zipf.read(filename)) > 0, f"Пустой или битый файл {filename}"

# Проверка на типы файлов

def check_files_types(archive_path):
    with zipfile.ZipFile(archive_path, 'r') as zipf:
        for filename in zipf.namelist():
            assert filename.endswith((".pdf", ".csv", ".xlsx")), f"Ошибка в файле {filename}"

# Проверка информации в файле pdf

def check_pdf(archive_path):
    with zipfile.ZipFile(archive_path, 'r') as zipf:
        for filename in zipf.namelist():
            if filename.endswith(".pdf"):
                with zipf.open(filename) as file:
                    reader = PdfReader(file)
                    assert "Python Testing with pytest" in reader.pages[1].extract_text(), "Текст в pdf не найден"


# Проверка информации в файле csv

def check_csv(archive_path):
    with zipfile.ZipFile(archive_path, 'r') as zipf:
        for filename in zipf.namelist():
            if filename.endswith(".csv"):
                with zipf.open(filename) as file:
                    file_content_str = file.read().decode('utf-8-sig')
                    reader = csv.reader(io.StringIO(file_content_str), delimiter=';')
                    rows = list(reader)
                    assert rows[4][2] == "1_1", "Текст в ячейке csv не соответствует"


# Проверка информации в файле xlsx

def check_xlsx(archive_path):
    with zipfile.ZipFile(archive_path, 'r') as zipf:
        for filename in zipf.namelist():
            if filename.endswith(".xlsx"):
                with zipf.open(filename) as file:
                    workbook = load_workbook(file)
                    sheet = workbook.active
                    assert sheet.cell(row=2, column=5).value == "Соответствует", "Текст в ячейке xlsx не соответствует"


archive_exists(archive_path)
amount_of_files(archive_path)
unic_files(archive_path)
files_are_not_empty(archive_path)
check_files_types(archive_path)
check_pdf(archive_path)
check_csv(archive_path)
check_xlsx(archive_path)

